import numpy as np
from unicodedata import normalize
from math import sqrt
from scipy import stats
from sklearn import linear_model, feature_selection, cluster
from statsmodels.tsa.statespace.sarimax import SARIMAX
from matplotlib import pyplot as plt
from consolemenu.format import MenuBorderStyleType
from consolemenu import SelectionMenu, MenuFormatBuilder
import statistics
import openpyxl
import io
import re

class LinearModel:
    def __init__(self, y, x = None):
        if x is not None:
            self.x = np.array(x)
        else:
            self.x = np.arange(1, len(y) + 1)
        self.reshaped_x = self.x.reshape(-1, 1)
        self.y = np.array(y)
        self.model = linear_model.LinearRegression()
        self.model.fit(self.reshaped_x, self.y)
        self.b = self.model.intercept_
        self.k = self.model.coef_[0]
        self.r2 = self.model.score(self.reshaped_x, self.y)
        self.r = sqrt(abs(self.r2))
        self.f = feature_selection.f_regression(self.reshaped_x, self.y)[0][0]
        self.f_t = stats.f.isf(0.05, 1, len(y) - 2)
        self.average_absolute_increase = np.average([y[i] - y[i - 1] for i in range(1, len(y))])
        self.average_growth_rate = np.exp(np.log([y[i] / y[i - 1] * 100 for i in range(1, len(y))]).mean())
        self.confidence_interval = self.__get_confidence_interval()

    def __get_confidence_interval(self):
        e = self.y - self.model.predict(self.reshaped_x)
        sse = np.sum(e ** 2)
        if (df := len(self.reshaped_x) - 2) == 0:
            return [None, None]
        mse = sse / df
        _x = np.average(self.reshaped_x)
        if (sum_diff := np.sum((self.reshaped_x - _x) ** 2)) == 0:
            return [None, None]
        se = sqrt(mse / sum_diff)
        t_t = stats.t.ppf(1 - 0.05 / 2, df)
        k = self.model.coef_[0]
        return [k - se * t_t, k + se * t_t]

class Program:
    def __init__(self):
        self.data_book = {}
        self.current_region = None
        self.regions = []
        self.current_sdg = None
        self.str_re = re.compile(r"[^а-яА-ЯЁё%.,\-() ]")
        self.menu_format = MenuFormatBuilder() \
            .set_border_style_type(MenuBorderStyleType.ASCII_BORDER) \
            .set_title_align('center') \
            .set_items_top_padding(0) \
            .set_items_bottom_padding(0) \
            .set_items_left_padding(0) \
            .set_items_right_padding(0) \
            .set_subtitle_align('center')
        self.load_data()
        self.show_regions()

    def load_data(self):
        print("Загрузка таблицы...")
        workbook = openpyxl.load_workbook("data.xlsx")
        for sheet in workbook.worksheets:
            titles = []
            reset_title = False
            last_labels = []
            for row in range(1, sheet.max_row + 1):
                first_column = normalize("NFKC", f"{sheet.cell(row, 1).value or ""}")
                first_column = self.str_re.sub("", first_column).strip()
                values = []
                labels = []
                count_no_empty_values = 0
                for column in range(2, sheet.max_column + 1):
                    value = f"{sheet.cell(row, column).value or ""}"
                    value = re.compile("[^0-9,.]").sub("", value).replace(",", ".").strip()
                    if first_column:
                        if value == "":
                            break
                        try:
                            values.append(float(value))
                            count_no_empty_values += 1
                        except ValueError:
                            values.append(None)
                    else:
                        try:
                            labels.append(int(value[:4]))
                        except ValueError:
                            break
                if len(labels) > 0:
                    last_labels = labels
                if count_no_empty_values >= 3 and len(values) == len(last_labels):
                    if not first_column in self.data_book:
                        self.data_book[first_column] = {}
                    title = "\n".join(titles[-2:])
                    self.data_book[first_column][title] = [last_labels, values]
                    reset_title = True
                elif not first_column == "":
                    if reset_title:
                        titles = []
                        reset_title = False
                    titles.append(first_column)
        try:
            with io.open("regions.txt", mode = "r", encoding = "utf-8") as file:
                regions = file.readlines()
                regions = [self.str_re.sub("", region).strip() for region in regions]
                self.regions = regions
        except:
            pass

    def show_regions(self, pair = False, all = False):
        regions = list(self.data_book.keys())
        if len(self.regions) > 0:
            regions = [region for region in regions if (not all and region in self.regions) or (all and region not in self.regions)]
        if pair:
            searched_data = self.searched_data.keys()
            regions = [region for region in regions if region in searched_data]
        if all:
            regions += ["Выбрать регион"]
        else:
            regions += ["Выбрать другой показатель"]
        menu = SelectionMenu(regions, "Выберите данные" if all else "Выберите регион", f"Парная регрессия: {"да" if pair else "нет"}", exit_option_text = "Выйти", formatter = self.menu_format)
        try:
            menu.show()
        except KeyboardInterrupt:
            print("Выход...")
            exit()
        if menu.selected_option == len(regions) - 1:
            self.show_regions(pair, not all)
            return
        if menu.selected_option == len(regions):
            if pair:
                self.show_regions()
            return
        self.show_sdgs(regions[menu.selected_option], pair)

    def show_sdgs(self, region, pair = False):
        sdgs = list(self.data_book[region].keys())
        if pair:
            searched_data = self.searched_data[region].keys()
            sdgs = [sdg for sdg in sdgs if sdg in searched_data]
        menu = SelectionMenu(sdgs, "Выберите показатель", f"{region}\nПарная регрессия: {"да" if pair else "нет"}", exit_option_text = "Выйти", formatter = self.menu_format)
        try:
            menu.show()
        except KeyboardInterrupt:
            print("Выход...")
            exit()
        if menu.selected_option <= len(sdgs) - 1:
            sdg = sdgs[menu.selected_option]
            self.show_menu(region, sdg)
        else:
            self.show_regions(pair)

    def search_pairs(self):
        self.searched_data = {}
        for region, sdgs in self.data_book.items():
            for sdg, data in sdgs.items():
                data_pair = self.data_book[self.current_region][self.current_sdg]
                prepared_data = self.prepare_data(data_pair[1], data_pair[0], data[1], data[0])
                model = LinearModel(self.apply_iqr(prepared_data[0]), self.apply_iqr(prepared_data[2]))
                if model.r >= 0.7 and model.f >= model.f_t:
                    if not region in self.searched_data:
                        self.searched_data[region] = {}
                    self.searched_data[region][sdg] = [model.r, model.f]
        return self.searched_data

    def show_cluster(self):
        kmeans = cluster.KMeans(n_clusters = 3)
        labels = []
        data_cluster = []
        for region, sdgs in self.searched_data.items():
            if region == self.current_region or region not in self.regions:
                for sdg, data in sdgs.items():
                    if not sdg == self.current_sdg:
                        data_cluster.append(data)
                        labels.append(f"{region}\n{sdg}")
        data_cluster = np.array(data_cluster)
        kmeans.fit(data_cluster)
        fig, ax = plt.subplots(figsize = (5, 3))
        ax.scatter(data_cluster[:, 0], data_cluster[:, 1], c = kmeans.labels_)
        for i, (x, y) in enumerate(data_cluster):
            ax.text(x, y, labels[i], fontsize = 7)
        ax.set_xlabel("Множественный R")
        ax.set_ylabel("Критерий Фишера")
        fig.tight_layout()
        fig.canvas.manager.set_window_title("Кластеризация показателей")
        try:
            plt.show()
        except KeyboardInterrupt:
            print("Выход...")
            exit()

    def show_menu(self, region, sdg):
        data_y = self.data_book[region][sdg]
        menu_items = [
            "Вернуться назад",
            "Рассчитать линейную регрессию",
            "Ликвидировать выбросы и рассчитать"
        ]
        if self.current_sdg is None:
            menu_items += [
                "Связать с другим показателем",
                "Найти парную регрессию",
                "Кластеризация показателей",
                "Прогноз ARIMA"
            ]
        menu = SelectionMenu(menu_items, "Выберите действия", f"{region}\n{sdg}", show_exit_option = False, formatter = self.menu_format)
        try:
            menu.show()
        except KeyboardInterrupt:
            print("Выход...")
            exit()
        data = self.prepare_data(data_y[1], data_y[0], data_y[1], data_y[0])
        if menu.selected_option == 0:
            self.current_region = self.current_sdg = None
            self.show_regions()
            return
        if menu.selected_option == 3:
            self.current_region = region
            self.current_sdg = sdg
            self.show_regions()
            return
        if menu.selected_option == 4:
            self.current_region = region
            self.current_sdg = sdg
            self.search_pairs()
            self.show_regions(True)
            return
        print("Закройте все окна с графиками")
        if menu.selected_option == 5:
            self.current_region = region
            self.current_sdg = sdg
            self.search_pairs()
            self.show_cluster()
            self.current_region = self.current_sdg = None
        elif menu.selected_option == 6:
            self.show_arima_model(data_y[1], data_y[0], f"{region}\n{sdg}")
        elif self.current_sdg is not None:
            data_x = self.data_book[self.current_region][self.current_sdg]
            data = self.prepare_data(data_y[1], data_y[0], data_x[1], data_x[0])
            if menu.selected_option == 1:
                self.show_model(data[0], data[1], f"y – {region}\n{sdg}", data[2], f"x – {self.current_region}\n{self.current_sdg}")
            elif menu.selected_option == 2:
                self.show_iqr_model(data[0], data[1], f"y – {region}\n{sdg}", data[2], f"x – {self.current_region}\n{self.current_sdg}")
        else:
            if menu.selected_option == 1:
                self.show_model(data[0], data[1], f"{region}\n{sdg}")
            elif menu.selected_option == 2:
                self.show_iqr_model(data[0], data[1], f"{region}\n{sdg}")
        self.show_menu(region, sdg)

    def prepare_data(self, y, y_labels, x, x_labels):
        filtered_y = []
        filtered_x = []
        filtered_labels = []
        for i, value in enumerate(y):
            label = y_labels[i]
            try:
                x_index = x_labels.index(label)
            except ValueError:
                continue
            if value is not None and x[x_index] is not None:
                filtered_y.append(value)
                filtered_x.append(x[x_index])
                filtered_labels.append(y_labels[i])
        return [filtered_y, filtered_labels, filtered_x]

    def apply_iqr(self, data):
        q1, q2, q3 = statistics.quantiles(data)
        q_min, q_max = [q1 - 1.5 * (q3 - q1), q3 + 1.5 * (q3 - q1)]
        return [q2 if el < q_min or el > q_max else el for el in data]

    def show_iqr_model(self, y, labels = None, y_title = None, x = None, x_title = None):
        pair = x is not None
        if pair:
            fig, axs = plt.subplots(1, 2, figsize = (3, 3))
            axs[1].set_title("x")
            axs[1].boxplot(x, patch_artist = True)
            axs[1].grid()
            ax = axs[0]
        else:
            fig, ax = plt.subplots(figsize = (2, 3))
        ax.set_title("y")
        ax.boxplot(y, patch_artist = True)
        ax.grid()
        fig.tight_layout()
        fig.canvas.manager.set_window_title("Диаграмма размаха")
        self.show_model(self.apply_iqr(y), labels, y_title, None if x is None else self.apply_iqr(x), x_title)

    def show_arima_model(self, y, labels, y_title = "y"):
        try:
            model = SARIMAX(y, order = (3, 1, 0), initialization = "approximate_diffuse")
            model_fit = model.fit(disp = 0)
        except:
            return
        st_pred = model_fit.get_prediction()
        next_pred = model_fit.get_forecast(steps = 3)
        mse = ((st_pred.predicted_mean[1:] - y[1:]) ** 2).mean()
        fig, ax = plt.subplots(figsize = (5, 3))
        ax.plot(y, label = "Данные")
        ax.plot(np.arange(1, len(y) + 3), np.concatenate((st_pred.predicted_mean[1:], next_pred.predicted_mean)), label = "Прогноз")
        ax.fill_between(np.arange(len(y), len(y) + 3), next_pred.conf_int()[:, 0], next_pred.conf_int()[:, 1], color = "k", alpha = 0.2)
        labels = np.array(labels)
        extended_labels = np.concatenate((labels, labels[-3:] + 3))
        ax.set_xticks(np.arange(0, len(y) + 3), extended_labels)
        ax.set_xticklabels(extended_labels)
        ax.set_title(y_title, fontsize = 9, wrap = True)
        ax.legend()
        fig.tight_layout()
        fig.canvas.manager.set_window_title(f"ARIMA MSE = {mse}")
        try:
            plt.show()
        except KeyboardInterrupt:
            print("Выход...")
            exit()

    def show_model(self, y, labels = None, y_title = "y", x = None, x_title = "x"):
        models = [LinearModel(y)]
        if x is not None:
            models.append(LinearModel(x))
        pair = len(models) == 2

        fig, axs = plt.subplots(2, 1, figsize = (5, 5))
        labels = np.array(labels)
        extended_labels = np.concatenate((labels, labels[-3:] + 3))
        for i, model in enumerate(models):
            extended_x = np.concatenate((model.x, model.x[-3:] + 3))
            axs[i].set_title(y_title if i == 0 else x_title, fontsize = 9, wrap = True)
            axs[i].plot(model.x, model.y)
            axs[i].plot(extended_x, model.k * extended_x + model.b)
            axs[i].set_xticks(extended_x, extended_labels)
            axs[i].set_xticklabels(extended_labels, rotation = "vertical")
            axs[i].grid()
        fig.tight_layout()
        fig.canvas.manager.set_window_title("Линейная регрессия")

        model = LinearModel(y, x) if pair else models[0]
        if pair:
            fig, ax = plt.subplots(figsize = (4, 2))
        else:
            ax = axs[1]
        table = ax.table([
            ["k-коэффициент", model.k],
            ["Свободный коэффициент", model.b],
            ["Множественный R", model.r],
            ["R-квадрат", model.r2],
            ["Критерий Фишера", model.f],
            ["F-табличное", model.f_t],
            ["Довер. интервал", model.confidence_interval],
            ["Средний абсолютный прирост", model.average_absolute_increase],
            ["Средний темп роста", f"{model.average_growth_rate}%"]
        ], cellLoc = "left", bbox = [0, 0, 1, 1], loc = "center")
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        ax.axis("off")

        if pair:
            fig.tight_layout()
            fig.canvas.manager.set_window_title("Параметрические данные")

        try:
            plt.show()
        except KeyboardInterrupt:
            print("Выход...")
            exit()

Program()
